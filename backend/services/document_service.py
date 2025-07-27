import os
import shutil
import mimetypes
from typing import List, Optional, Dict, Any
from pathlib import Path
from datetime import datetime
import PyPDF2
import docx
import openpyxl
import json
import asyncio
import uuid
from motor.motor_asyncio import AsyncIOMotorClient

from models.document import DocumentMetadata, DocumentSearch, DocumentSearchResult
from services.llm_service import LLMService
from services.embedding_service import EmbeddingService
from utils.json_encoder import serialize_document, serialize_documents

class DocumentService:
    def __init__(self, db: AsyncIOMotorClient, storage_path: str = "/app/backend/storage"):
        self.db = db
        self.storage_path = Path(storage_path)
        self.storage_path.mkdir(exist_ok=True)
        self.llm_service = LLMService()
        self.embedding_service = EmbeddingService()
        
    async def upload_document(self, file_data: bytes, filename: str, user_id: str, 
                            category: str = "general", tags: List[str] = None, 
                            auto_categorize: bool = True) -> DocumentMetadata:
        """Upload and process a document"""
        try:
            # Generate unique filename
            doc_id = str(uuid.uuid4())
            file_ext = Path(filename).suffix.lower()
            stored_filename = f"{doc_id}{file_ext}"
            file_path = self.storage_path / stored_filename
            
            # Save file to storage
            with open(file_path, 'wb') as f:
                f.write(file_data)
            
            # Extract metadata
            file_size = len(file_data)
            mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
            
            # Extract text content
            extracted_text = await self._extract_text(file_path, mime_type)
            
            # Auto-categorize if requested
            if auto_categorize and extracted_text:
                category = await self._auto_categorize(extracted_text, filename)
            
            # Generate tags if not provided
            if not tags and extracted_text:
                tags = await self._generate_tags(extracted_text, filename)
            
            # Generate embeddings
            embedding = await self.embedding_service.generate_embedding(extracted_text or filename)
            
            # Create document metadata
            document = DocumentMetadata(
                id=doc_id,
                filename=stored_filename,
                original_filename=filename,
                file_path=str(file_path),
                file_size=file_size,
                file_type=file_ext.lstrip('.'),
                mime_type=mime_type,
                category=category,
                tags=tags or [],
                user_id=user_id,
                extracted_text=extracted_text,
                embedding=embedding,
                content_summary=await self._generate_summary(extracted_text) if extracted_text else None
            )
            
            # Store in database
            document_dict = document.dict()
            await self.db.documents.insert_one(document_dict)
            
            return document
            
        except Exception as e:
            # Clean up file if something went wrong
            if file_path.exists():
                file_path.unlink()
            raise Exception(f"Failed to upload document: {str(e)}")
    
    async def _extract_text(self, file_path: Path, mime_type: str) -> Optional[str]:
        """Extract text content from various file types"""
        try:
            if mime_type == "application/pdf":
                return await self._extract_pdf_text(file_path)
            elif mime_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
                              "application/msword"]:
                return await self._extract_docx_text(file_path)
            elif mime_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              "application/vnd.ms-excel"]:
                return await self._extract_excel_text(file_path)
            elif mime_type.startswith("text/"):
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            else:
                return None
        except Exception as e:
            print(f"Error extracting text from {file_path}: {e}")
            return None
    
    async def _extract_pdf_text(self, file_path: Path) -> str:
        """Extract text from PDF files"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            print(f"Error extracting PDF text: {e}")
        return text.strip()
    
    async def _extract_docx_text(self, file_path: Path) -> str:
        """Extract text from Word documents"""
        try:
            doc = docx.Document(file_path)
            text = []
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
            return "\n".join(text)
        except Exception as e:
            print(f"Error extracting DOCX text: {e}")
            return ""
    
    async def _extract_excel_text(self, file_path: Path) -> str:
        """Extract text from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) for cell in row if cell is not None]
                    if row_text:
                        text.append(" ".join(row_text))
            return "\n".join(text)
        except Exception as e:
            print(f"Error extracting Excel text: {e}")
            return ""
    
    async def _auto_categorize(self, text: str, filename: str) -> str:
        """Auto-categorize document based on content"""
        # Mock categorization logic - replace with LLM call
        filename_lower = filename.lower()
        text_lower = text.lower()
        
        if any(word in filename_lower or word in text_lower for word in ['contract', 'agreement', 'terms']):
            return 'contracts'
        elif any(word in filename_lower or word in text_lower for word in ['invoice', 'bill', 'payment']):
            return 'invoices'
        elif any(word in filename_lower or word in text_lower for word in ['hr', 'employee', 'onboarding', 'hire']):
            return 'hr'
        elif any(word in filename_lower or word in text_lower for word in ['compliance', 'policy', 'regulation']):
            return 'compliance'
        else:
            return 'general'
    
    async def _generate_tags(self, text: str, filename: str) -> List[str]:
        """Generate tags for document"""
        # Mock tag generation - replace with LLM call
        tags = []
        text_lower = text.lower()
        filename_lower = filename.lower()
        
        # Common business document tags
        tag_keywords = {
            'urgent': ['urgent', 'asap', 'immediate', 'priority'],
            'signed': ['signed', 'signature', 'executed'],
            'draft': ['draft', 'preliminary', 'version'],
            'quarterly': ['quarterly', 'q1', 'q2', 'q3', 'q4'],
            'annual': ['annual', 'yearly', 'year-end'],
            'confidential': ['confidential', 'private', 'restricted'],
            'review': ['review', 'approval', 'pending'],
        }
        
        for tag, keywords in tag_keywords.items():
            if any(keyword in text_lower or keyword in filename_lower for keyword in keywords):
                tags.append(tag)
        
        return tags[:5]  # Limit to 5 tags
    
    async def _generate_summary(self, text: str) -> Optional[str]:
        """Generate document summary"""
        if not text or len(text) < 100:
            return None
        
        # Mock summary generation - replace with LLM call
        return f"Document summary: {text[:200]}..."
    
    async def search_documents(self, search: DocumentSearch, user_id: str) -> List[DocumentSearchResult]:
        """Search documents using text and embedding similarity"""
        try:
            # Build MongoDB query
            query = {"user_id": user_id}
            
            if search.categories:
                query["category"] = {"$in": search.categories}
            
            if search.tags:
                query["tags"] = {"$in": search.tags}
            
            # Text search
            if search.query:
                query["$or"] = [
                    {"original_filename": {"$regex": search.query, "$options": "i"}},
                    {"extracted_text": {"$regex": search.query, "$options": "i"}},
                    {"content_summary": {"$regex": search.query, "$options": "i"}}
                ]
            
            # Find documents
            documents = await self.db.documents.find(query).limit(search.limit).to_list(search.limit)
            
            results = []
            for doc in documents:
                # Serialize document
                doc = serialize_document(doc)
                
                # Calculate mock relevance score
                relevance_score = await self._calculate_relevance_score(doc, search.query)
                
                document_obj = DocumentMetadata(**doc)
                
                # Extract matching content if requested
                matching_content = None
                if search.include_content and search.query and doc.get('extracted_text'):
                    matching_content = await self._extract_matching_content(doc['extracted_text'], search.query)
                
                results.append(DocumentSearchResult(
                    document=document_obj,
                    relevance_score=relevance_score,
                    matching_content=matching_content
                ))
            
            # Sort by relevance score
            results.sort(key=lambda x: x.relevance_score, reverse=True)
            
            return results
            
        except Exception as e:
            print(f"Error searching documents: {e}")
            return []
    
    async def _calculate_relevance_score(self, document: dict, query: str) -> float:
        """Calculate relevance score for search results"""
        if not query:
            return 0.5
        
        score = 0.0
        query_lower = query.lower()
        
        # Check filename match
        if query_lower in document.get('original_filename', '').lower():
            score += 0.3
        
        # Check content match
        if document.get('extracted_text') and query_lower in document['extracted_text'].lower():
            score += 0.4
        
        # Check tags match
        if any(query_lower in tag.lower() for tag in document.get('tags', [])):
            score += 0.2
        
        # Check category match
        if query_lower in document.get('category', '').lower():
            score += 0.1
        
        return min(score, 1.0)
    
    async def _extract_matching_content(self, text: str, query: str) -> str:
        """Extract content snippet that matches the query"""
        if not text or not query:
            return ""
        
        query_lower = query.lower()
        text_lower = text.lower()
        
        # Find the position of the query in the text
        pos = text_lower.find(query_lower)
        if pos == -1:
            return text[:200] + "..." if len(text) > 200 else text
        
        # Extract context around the match
        start = max(0, pos - 100)
        end = min(len(text), pos + len(query) + 100)
        
        snippet = text[start:end]
        if start > 0:
            snippet = "..." + snippet
        if end < len(text):
            snippet = snippet + "..."
        
        return snippet
    
    async def get_document_by_id(self, doc_id: str, user_id: str) -> Optional[DocumentMetadata]:
        """Get document by ID"""
        try:
            doc = await self.db.documents.find_one({"id": doc_id, "user_id": user_id})
            if doc:
                doc = serialize_document(doc)
                return DocumentMetadata(**doc)
            return None
        except Exception as e:
            print(f"Error getting document: {e}")
            return None
    
    async def get_user_documents(self, user_id: str, limit: int = 100) -> List[DocumentMetadata]:
        """Get all documents for a user"""
        try:
            documents = await self.db.documents.find({"user_id": user_id}).limit(limit).to_list(limit)
            documents = serialize_documents(documents)
            return [DocumentMetadata(**doc) for doc in documents]
        except Exception as e:
            print(f"Error getting user documents: {e}")
            return []
    
    async def delete_document(self, doc_id: str, user_id: str) -> bool:
        """Delete a document"""
        try:
            doc = await self.db.documents.find_one({"id": doc_id, "user_id": user_id})
            if not doc:
                return False
            
            # Delete file from storage
            file_path = Path(doc['file_path'])
            if file_path.exists():
                file_path.unlink()
            
            # Delete from database
            result = await self.db.documents.delete_one({"id": doc_id, "user_id": user_id})
            return result.deleted_count > 0
            
        except Exception as e:
            print(f"Error deleting document: {e}")
            return False